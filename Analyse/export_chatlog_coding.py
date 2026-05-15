"""
export_chatlog_coding.py — Eksporterer chatlogge til Excel til manuel kodning.

Struktur: én række per besked, i kronologisk rækkefølge.
Efter hver deltagers beskeder indsættes en gennemsnitsrække med AVERAGEIF-formler:
  - CHAT-SUM + CHAT-SCAF: gennemsnit af chatbot-rækker
  - PERS-SUM + PERS-SCAF-RESP: gennemsnit af deltager-rækker

Kolonner: ParticipantID | Aktør | Ytring | CHAT-SUM | PERS-SUM | CHAT-SCAF | PERS-SCAF-RESP

Brug: python export_chatlog_coding.py
"""

import json
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.chdir(os.path.dirname(os.path.abspath(__file__)))

# ── Farver ────────────────────────────────────────────────────────────────────
CLR_HEADER     = "1A202C"   # meget mørk (header)
CLR_CHATBOT    = "EBF5FF"   # lyseblå — chatbot-rækker
CLR_DELTAGER   = "F0FFF4"   # lysegrøn — deltager-rækker
CLR_SUMMARY    = "FFF9E6"   # lys gul — gennemsnitsrækker
CLR_SEPARATOR  = "F7F7F7"   # meget lys grå — tomme separatorrækker
CLR_WHITE      = "FFFFFF"

THIN = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_messages(transcript_str):
    if pd.isna(transcript_str):
        return []
    try:
        msgs = json.loads(transcript_str)
        return [(m.get("role", ""), m.get("content", "").strip()) for m in msgs]
    except Exception:
        return []


def fill(cell, hex_color):
    cell.fill = PatternFill("solid", fgColor=hex_color)


def run():
    df = pd.read_csv("data/processed/processed.csv")
    df = df[df["chatTranscript"].notna()].copy()
    df["createdAt"] = pd.to_datetime(df["createdAt"], utc=True)
    df = df.sort_values("createdAt").reset_index(drop=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Kodning"

    # ── Header ────────────────────────────────────────────────────────────────
    headers = ["#", "ParticipantID", "Aktør", "Ytring",
               "CHAT-SUM", "PERS-SUM", "CHAT-SCAF", "PERS-SCAF-RESP"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=CLR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
    ws.row_dimensions[1].height = 30

    # ── Kolonner A-H ──────────────────────────────────────────────────────────
    #  A=#, B=ParticipantID, C=Aktør, D=Ytring, E=CHAT-SUM, F=PERS-SUM,
    #  G=CHAT-SCAF, H=PERS-SCAF-RESP
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 70
    for col in ["E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 14

    # Frys øverste række
    ws.freeze_panes = "A2"

    current_row = 2

    for participant_nr, (_, participant) in enumerate(df.iterrows(), start=1):
        pid   = participant["participantId"]
        msgs  = parse_messages(participant["chatTranscript"])
        if not msgs:
            continue

        block_start = current_row

        for i, (role, content) in enumerate(msgs):
            is_chatbot = role == "assistant"
            aktor      = "Chatbot" if is_chatbot else "Deltager"
            bg         = CLR_CHATBOT if is_chatbot else CLR_DELTAGER

            ws.cell(current_row, 1, participant_nr if i == 0 else "")
            ws.cell(current_row, 2, pid if i == 0 else "")
            ws.cell(current_row, 3, aktor)
            ws.cell(current_row, 4, content)
            # Tomme kodningsfelter (E-H)
            for col in range(5, 9):
                ws.cell(current_row, col, "")

            for col in range(1, 9):
                cell = ws.cell(current_row, col)
                fill(cell, bg)
                cell.border    = THIN_BORDER
                cell.alignment = Alignment(vertical="top",
                                           wrap_text=(col == 4))

            ws.row_dimensions[current_row].height = 15
            current_row += 1

        block_end = current_row - 1

        # ── Gennemsnitsrække med AVERAGEIF-formler ────────────────────────────
        # CHAT-SUM og CHAT-SCAF: kun chatbot-rækker (C = "Chatbot")
        # PERS-SUM og PERS-SCAF-RESP: kun deltager-rækker (C = "Deltager")
        c_range = f"C{block_start}:C{block_end}"
        ws.cell(current_row, 1, participant_nr)
        ws.cell(current_row, 2, pid)
        ws.cell(current_row, 3, "GENNEMSNIT")
        ws.cell(current_row, 4, f"↑ {len(msgs)} beskeder  |  "
                                f"chatbot: {sum(1 for r,_ in msgs if r=='assistant')}  |  "
                                f"deltager: {sum(1 for r,_ in msgs if r=='user')}")
        ws.cell(current_row, 5,
                f'=IFERROR(AVERAGEIF({c_range},"Chatbot",E{block_start}:E{block_end}),"")')
        ws.cell(current_row, 6,
                f'=IFERROR(AVERAGEIF({c_range},"Deltager",F{block_start}:F{block_end}),"")')
        ws.cell(current_row, 7,
                f'=IFERROR(AVERAGEIF({c_range},"Chatbot",G{block_start}:G{block_end}),"")')
        ws.cell(current_row, 8,
                f'=IFERROR(AVERAGEIF({c_range},"Deltager",H{block_start}:H{block_end}),"")')

        for col in range(1, 9):
            cell = ws.cell(current_row, col)
            fill(cell, CLR_SUMMARY)
            cell.border    = THIN_BORDER
            cell.font      = Font(bold=True, italic=(col in [5,6,7,8]))
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if col >= 5 else "left")
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # ── Tom separatorlinje ────────────────────────────────────────────────
        for col in range(1, 9):
            cell = ws.cell(current_row, col)
            fill(cell, CLR_SEPARATOR)
        ws.row_dimensions[current_row].height = 8
        current_row += 1

    # ── Skalaforklaring i bunden ──────────────────────────────────────────────
    current_row += 1
    note = ("Skala: 0 = slet ikke opfyldt  |  1 = delvist eller svagt opfyldt  "
            "|  2 = tydeligt eller stærkt opfyldt")
    ws.cell(current_row, 1, note)
    ws.cell(current_row, 1).font = Font(italic=True, color="888888", size=9)

    current_row += 1
    ws.cell(current_row, 1,
            "NB: CHAT-SUM + CHAT-SCAF kodes kun på chatbot-rækker  |  "
            "PERS-SUM + PERS-SCAF-RESP kodes kun på deltager-rækker  |  "
            "Skriv 0 på de felter der ikke er relevante for aktøren")
    ws.cell(current_row, 1).font = Font(italic=True, color="888888", size=9)

    out_path = "data/processed/chatlog_kodning_blank.xlsx"
    wb.save(out_path)

    n_ctrl = (df["group"] == "control").sum()
    n_intr = (df["group"] == "intervention").sum()
    total_msgs = sum(len(parse_messages(r["chatTranscript"])) for _, r in df.iterrows())
    print(f"  Gemt → {out_path}")
    print(f"  {len(df)} deltagere  (kontrol={n_ctrl}, intervention={n_intr})")
    print(f"  {total_msgs} rækker i alt (én per besked)")
    print(f"  Gennemsnitsrækker beregner automatisk når du udfylder kodningen")


if __name__ == "__main__":
    run()
